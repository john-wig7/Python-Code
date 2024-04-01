import pandas as pd
from openpyxl import load_workbook
import requests
from io import BytesIO

file_path = "https://onedrive.live.com/download?resid=7DD412FE3A3475E2%21160296&authkey=!ACznhiop5FP1t5E&em=2"

response = requests.get(file_path)
excel_data = response.content


wb = load_workbook(filename=BytesIO(excel_data))

sheet_name = 'ReceiptSheet'
ws = wb[sheet_name]

table_name = 'ReceiptsTable'
data = ws[table_name].values
columns = next(data)

df = pd.DataFrame(data, columns=columns)

wb.close

print (df.head())